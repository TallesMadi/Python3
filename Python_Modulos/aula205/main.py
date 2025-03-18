from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER /  'workbook.xlsx'

workbook = Workbook()
worksheet: Worksheet = workbook.active #type: ignore

worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')

students = [
    ['Talles', 20, 10],
    ['Raquel', 22, 10],
    ['Luna', 3, 7],
    ['Tralalero Tralala', 50, 1],
]

# for i, student_row in enumerate(students, start=2):
#     for j, student_column in enumerate(student_row, start=1):
#         worksheet.cell(i, j, student_column)

for student in students:
    worksheet.append(student)

workbook.save(WORKBOOK_PATH)